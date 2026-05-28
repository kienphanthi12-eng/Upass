import sys
import io
import os
from pathlib import Path

# Fix terminal encoding issues
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

sys.path.insert(0, str(Path(__file__).parent.parent))

xlsx_path = r"C:\Users\HP\Downloads\cấu trúc đề thi thpt tất cả các môn.xlsx"

def main():
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        return
        
    print(f"Reading file: {xlsx_path}")
    try:
        import pandas as pd
        excel = pd.ExcelFile(xlsx_path)
        print("Sheet names:", excel.sheet_names)
        for sheet_name in excel.sheet_names:
            df = excel.parse(sheet_name)
            print(f"\n--- Sheet: {sheet_name} ---")
            print(df.head(20))
    except ImportError:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            print("Sheet names:", wb.sheetnames)
            for name in wb.sheetnames:
                print(f"\n--- Sheet: {name} ---")
                ws = wb[name]
                for r in list(ws.iter_rows(values_only=True))[:20]:
                    print(r)
        except ImportError:
            print("Both pandas and openpyxl are missing. Let's try installing openpyxl or pandas using pip.")
            import subprocess
            subprocess.run(["pip", "install", "openpyxl"])
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            print("Sheet names:", wb.sheetnames)
            for name in wb.sheetnames:
                print(f"\n--- Sheet: {name} ---")
                ws = wb[name]
                for r in list(ws.iter_rows(values_only=True))[:20]:
                    print(r)

if __name__ == "__main__":
    main()
